import yfinance as yf
import pandas as pd
import os
from openpyxl import load_workbook
from datetime import datetime

def load_tickers(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".xlsx":
        df = pd.read_excel(file_path)
        tickers = df.iloc[:, 0].dropna().astype(str).tolist()
    elif ext == ".txt":
        with open(file_path, "r") as f:
            tickers = [line.strip() for line in f if line.strip()]
    else:
        raise ValueError("File must be .xlsx or .txt")
    return tickers

def calculate_cagr(start, end, periods):
    try:
        if start > 0 and end is not None:
            return (end / start) ** (1 / periods) - 1
    except:
        return None
    return None

def get_price_change(stock, months=None, years=None):
    try:
        if months:
            start = pd.Timestamp.now() - pd.DateOffset(months=months)
        elif years:
            start = pd.Timestamp.now() - pd.DateOffset(years=years)
        else:
            return None
        hist = stock.history(start=start, end=pd.Timestamp.now())
        if hist.empty:
            return None
        start_price = hist["Close"].iloc[0]
        end_price = hist["Close"].iloc[-1]
        return (end_price / start_price - 1) if start_price else None
    except:
        return None

def get_ncav(ticker):
    try:
        stock = yf.Ticker(ticker)
        bs = stock.balance_sheet
        info = stock.info
        income = stock.income_stmt

        current_assets = None
        total_liabilities = None
        ncav = None
        market_cap = info.get("marketCap", None)
        price_ncav_ratio = None
        netnet = False
        deep_netnet = False

        # Current Assets és Total Liabilities
        if not bs.empty:
            for idx in bs.index:
                idx_clean = idx.strip()
                if idx_clean == "Current Assets" and current_assets is None:
                    current_assets = bs.loc[idx].iloc[0]
                if "total liabilities" in idx.lower() and total_liabilities is None:
                    total_liabilities = bs.loc[idx].iloc[0]

        if current_assets is not None and total_liabilities is not None and market_cap is not None:
            ncav = current_assets - total_liabilities
            if ncav > 0:
                price_ncav_ratio = market_cap / ncav
                netnet = market_cap < ncav
                deep_netnet = market_cap < 0.67 * ncav

        # P/E
        pe_ratio = info.get("trailingPE", None)

        # Latest Price és Currency
        latest_price = info.get("regularMarketPrice", None)
        currency = info.get("currency", None)

        # TTM Dividend számítás timezone kezeléssel
        dividends = stock.dividends
        ttm_dividend = None
        if not dividends.empty:
            last_year = pd.Timestamp.now() - pd.DateOffset(years=1)
            dividends_naive = dividends.copy()
            if dividends_naive.index.tz is not None:
                dividends_naive.index = dividends_naive.index.tz_localize(None)
            ttm_dividend = dividends_naive[dividends_naive.index >= last_year].sum()

        dividend_rate = ttm_dividend
        if dividend_rate and latest_price:
            dividend_yield = dividend_rate / latest_price
        else:
            dividend_yield = None

        # EPS 5Y CAGR
        eps_cagr = None
        if income is not None and "Net Income" in income.index:
            net_income_series = income.loc["Net Income"].dropna()
            if len(net_income_series) >= 2:
                last_eps = net_income_series.iloc[0]
                first_eps = net_income_series.iloc[min(4, len(net_income_series)-1)]
                eps_cagr = calculate_cagr(first_eps, last_eps, min(5, len(net_income_series)-1))

        # Árfolyam változások
        price_1m = get_price_change(stock, months=1)
        price_3m = get_price_change(stock, months=3)
        price_6m = get_price_change(stock, months=6)
        price_1y = get_price_change(stock, years=1)
        price_3y = get_price_change(stock, years=3)
        price_5y = get_price_change(stock, years=5)

        return {
            "Ticker": ticker,
            "Current Assets": current_assets,
            "Total Liabilities": total_liabilities,
            "NCAV": ncav,
            "Market Cap": market_cap,
            "Price/NCAV": price_ncav_ratio,
            "Net-Net (<1)": netnet,
            "Deep Net-Net (<0.67)": deep_netnet,
            "P/E": pe_ratio,
            "Dividend Rate": dividend_rate,
            "Dividend Yield": dividend_yield,
            "EPS 5Y CAGR": eps_cagr,
            "Latest Price": latest_price,
            "Currency": currency,
            "Price Change 1M": price_1m,
            "Price Change 3M": price_3m,
            "Price Change 6M": price_6m,
            "Price Change 1Y": price_1y,
            "Price Change 3Y": price_3y,
            "Price Change 5Y": price_5y
        }

    except Exception as e:
        print(f"Error processing {ticker}: {e}")
        return {
            "Ticker": ticker,
            "Current Assets": None,
            "Total Liabilities": None,
            "NCAV": None,
            "Market Cap": None,
            "Price/NCAV": None,
            "Net-Net (<1)": False,
            "Deep Net-Net (<0.67)": False,
            "P/E": None,
            "Dividend Rate": None,
            "Dividend Yield": None,
            "EPS 5Y CAGR": None,
            "Latest Price": None,
            "Currency": None,
            "Price Change 1M": None,
            "Price Change 3M": None,
            "Price Change 6M": None,
            "Price Change 1Y": None,
            "Price Change 3Y": None,
            "Price Change 5Y": None
        }

def run_netnet_screener(input_file):
    tickers = load_tickers(input_file)
    print(f"Loaded {len(tickers)} tickers.")

    results = []
    for t in tickers:
        data = get_ncav(t)
        results.append(data)

    df = pd.DataFrame(results)

    # Időbélyeg hozzáadása a fájlnévhez
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    output_file = f"netnet_results_{timestamp}.xlsx"

    df.to_excel(output_file, index=False, engine="openpyxl")
    print(f"Results saved to {output_file}")

    # Excel formázás
    wb = load_workbook(output_file)
    ws = wb.active
    ws.freeze_panes = "A2"  # Első sor freeze
    header = {cell.value: cell.column_letter for cell in ws[1]}

    # Számformátum 1000-es szeparátorral
    num_cols = ["Current Assets", "Total Liabilities", "NCAV", "Market Cap"]
    for col_name in num_cols:
        if col_name in header:
            col_letter = header[col_name]
            for cell in ws[col_letter][1:]:
                if cell.value is not None:
                    cell.number_format = '#,##0'

    # 2 tizedesjegy (P/E, Dividend Rate, Latest Price)
    two_decimal_cols = ["P/E", "Dividend Rate", "Latest Price"]
    for col_name in two_decimal_cols:
        if col_name in header:
            col_letter = header[col_name]
            for cell in ws[col_letter][1:]:
                if cell.value is not None:
                    cell.number_format = '0.00'

    # Százalék formátum
    pct_cols = ["Price/NCAV", "Dividend Yield", "EPS 5Y CAGR",
                "Price Change 1M", "Price Change 3M", "Price Change 6M",
                "Price Change 1Y", "Price Change 3Y", "Price Change 5Y"]
    for col_name in pct_cols:
        if col_name in header:
            col_letter = header[col_name]
            for cell in ws[col_letter][1:]:
                if cell.value is not None:
                    cell.number_format = '0.00%'

    # Auto-fit column width (+ extra a Market Cap / NCAV nagy számokhoz)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                value_length = len(str(cell.value))
                if value_length > max_length:
                    max_length = value_length
            except:
                pass
        extra_width = 5 if col[0].value in ["Market Cap", "NCAV"] else 2
        ws.column_dimensions[column].width = max_length + extra_width

    wb.save(output_file)

    # Net-Net jelöltek
    netnet_candidates = df[df["Net-Net (<1)"] == True]
    deep_netnet_candidates = df[df["Deep Net-Net (<0.67)"] == True]

    if not netnet_candidates.empty:
        print("\nNet-Net Candidates (<1):")
        print(netnet_candidates[["Ticker", "Market Cap", "NCAV", "Price/NCAV"]])
    else:
        print("\nNo Net-Net candidates found. All tickers are marked False.")

    if not deep_netnet_candidates.empty:
        print("\nDeep Net-Net Candidates (<0.67):")
        print(deep_netnet_candidates[["Ticker", "Market Cap", "NCAV", "Price/NCAV"]])
    else:
        print("\nNo Deep Net-Net candidates found. All tickers are marked False.")

    return df, netnet_candidates, deep_netnet_candidates

if __name__ == "__main__":
    input_file = "tickers.txt"  # vagy "tickers.xlsx"
    run_netnet_screener(input_file)